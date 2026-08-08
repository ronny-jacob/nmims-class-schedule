import json, re, os
from collections import OrderedDict
from datetime import datetime, timedelta
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
def src(name):
    return os.path.join(BASE_DIR, "sources", name)
def resolve(p):
    if not p:
        return p
    return p if os.path.isabs(p) else os.path.join(BASE_DIR, p)

STUDENT_LIST = src("Division wise List- Trimester IV.xlsx")
LAST_YEAR_LIST = src("First Year Division list.xlsx")
TIMETABLE    = "downloads/03.08.2026 to 9.08.2026.xlsx"
TIMETABLE_NEXT = ""
FOOD_MENU    = "sources/August-Sept Menu Updated.xlsx"
FOOD_MENU_ANCHOR = "2026-08-03"
PLACEMENTS   = "sources/placements.json"
OUTPUT       = "data.json"
PLACEMENT_WINDOW_DAYS = 1

TIMETABLE = resolve(TIMETABLE)
TIMETABLE_NEXT = resolve(TIMETABLE_NEXT)
FOOD_MENU = resolve(FOOD_MENU)

FOOD_MEALS = [
    {"key": "breakfast", "label": "Breakfast", "time": "8:00 To 9:30"},
    {"key": "lunch",     "label": "Lunch",     "time": "12:30 To 2:30"},
    {"key": "snacks",    "label": "Snacks",    "time": "5:30 To 6:30"},
    {"key": "dinner",    "label": "Dinner",    "time": "8:00 To 9:30"},
]
FOOD_DAY_KEYS = {
    'MON': 'Mon', 'TUE': 'Tue', 'WED': 'Wed', 'THU': 'Thu',
    'FRI': 'Fri', 'SAT': 'Sat', 'SUN': 'Sun',
}

def parse_date_range(filepath):
    mt = re.search(r'(\d+)\.(\d+)\.(\d+)\s*to\s*(\d+)\.(\d+)\.(\d+)', filepath)
    if mt:
        d1, m1, y1, d2, m2, y2 = mt.groups()
        fmt = "%d.%m.%Y"
        dt_from = datetime.strptime(f"{d1}.{m1}.{y1}", fmt)
        dt_to   = datetime.strptime(f"{d2}.{m2}.{y2}", fmt)
        return dt_from.strftime("%a %-d %b") + " – " + dt_to.strftime("%a %-d %b %Y")
    return ""

def get_week_iso(filepath):
    mt = re.search(r'(\d+)\.(\d+)\.(\d+)\s*to\s*(\d+)\.(\d+)\.(\d+)', filepath)
    if mt:
        d1, m1, y1, d2, m2, y2 = mt.groups()
        fmt = "%d.%m.%Y"
        dt_from = datetime.strptime(f"{d1}.{m1}.{y1}", fmt)
        dt_to   = datetime.strptime(f"{d2}.{m2}.{y2}", fmt)
        return (dt_from.strftime("%Y-%m-%d"), dt_to.strftime("%Y-%m-%d"))
    return ("", "")

# ─── Subject name mappings ───────────────────────────────────────────────
SUBJECT_NAMES = {
    "AFSA_A": "Advanced Financial Statement Analysis A",
    "AFSA_B": "Advanced Financial Statement Analysis B",
    "BV_A":   "Business Valuation A",
    "BV_B":   "Business Valuation B",
    "IAPM_A": "Investment Analysis & Portfolio Management A",
    "IAPM_B": "Investment Analysis & Portfolio Management B",
    "CBM":    "Commercial Bank Management",
    "FD":     "Financial Derivative",
    "CB":     "Consumer Behaviour",
    "IMC":    "Integrated Marketing Communication",
    "MA":     "Marketing Analytics",
    "PRS":    "Pricing Strategy",
    "PS":     "Product Strategy",
    "MC_A":   "Management Consulting A",
    "MC_B":   "Management Consulting B",
    "ENT":    "Entrepreneurship",
    "LD":     "Learning & Development",
    "RS":     "Recruitment and Selection",
    "TR":     "Total Reward",
    "AI":     "Artificial Intelligence for Managers",
    "DAB":    "Data Analytics for Business",
    "BS":     "Business Simulation",
}

TIMETABLE_SUBJECT_MAP = {
    "MC-Div B":     ("MC_B", "Div B"),
    "MC-Div A":     ("MC_A", "Div A"),
    "MC_Div B":     ("MC_B", "Div B"),
    "MC_Div A":     ("MC_A", "Div A"),
    "Pricing Strategy": ("PRS", None),
    "AFSA-Div B":   ("AFSA_B", "Div B"),
    "AFSA-Div A":   ("AFSA_A", "Div A"),
    "AFSA_A":       ("AFSA_A", None),
    "AFSA_B":       ("AFSA_B", None),
    "L & D":        ("LD", None),
    "AIM":          ("AI", None),
    "R& S":         ("RS", None),
    "MA":           ("MA", None),
    "CBM":          ("CBM", None),
    "Integrated Marketing Communication": ("IMC", None),
    "Integrated Marketing": ("IMC", None),
    "CB":           ("CB", None),
    "DAB":          ("DAB", None),
    "Product Strategy": ("PS", None),
    "Financial Derivatives": ("FD", None),
    "BV_Div B":     ("BV_B", "Div B"),
    "BV_Div A":     ("BV_A", "Div A"),
    "IAPM_A":       ("IAPM_A", None),
    "IAPM_B":       ("IAPM_B", None),
    "BS Div A":     ("BS", "Div A"),
    "BS Div B":     ("BS", "Div B"),
    "BS Div C":     ("BS", "Div C"),
    "Total Rewards": ("TR", None),
    "Entrepreneurship": ("ENT", None),
}

def normalize(s, underscores=False):
    s = re.sub(r'\s+', ' ', s).strip()
    if underscores:
        s = re.sub(r'\s*_\s*', '_', s)
    return s

def parse_time_labels_from_sheet(ws):
    labels = OrderedDict()
    prev = ''
    for col_idx in range(2, ws.max_column + 1):
        cell = ws.cell(row=2, column=col_idx)
        if cell and cell.value:
            raw = normalize(str(cell.value))
            raw = re.sub(r'\s*[ap]m\s*', '', raw, flags=re.I)
            raw = re.sub(r'^0(\d)', r'\1', raw)
            raw = re.sub(r'(\s)0(\d)', r'\1\2', raw)
            raw = re.sub(r'\s*-\s*', '-', raw)
            labels[col_idx] = raw
            prev = raw
        elif prev:
            labels[col_idx] = prev
    return labels

def dedupe_consecutive(lst):
    if not lst:
        return lst
    result = [lst[0]]
    for item in lst[1:]:
        if item != result[-1]:
            result.append(item)
    return result

def match_subject(text, key):
    if text.startswith(key):
        return True
    if key.startswith(text):
        if len(text) == len(key):
            return True
        nxt = key[len(text)]
        return not nxt.isalnum()
    return False

def parse_timetable(filepath):
    wb = openpyxl.load_workbook(filepath)
    # Extract sheet name from the start date in filename
    mt = re.search(r'(\d+)\.(\d+)\.(\d+)\s*to', filepath)
    if mt:
        d, m, y = mt.groups()
        sheet_name = f"{int(d):02d}.{int(m):02d}.{y}"
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
    else:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    time_labels = parse_time_labels_from_sheet(ws)

    days_rows = [
        ("Mon", 3, 4), ("Tue", 5, 6), ("Wed", 7, 8),
        ("Thu", 9, 10), ("Fri", 11, 12), ("Sat", 13, 14), ("Sun", 15, 16)
    ]
    timetable = []

    for day_name, r1, r2 in days_rows:
        for row_idx in (r1, r2):
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                raw = cell.value
                if not raw:
                    continue
                raw_lines = [l.strip() for l in str(raw).split('\n') if l.strip()]
                # Skip decorative cells like vertical LUNCH BREAK text
                if all(len(l) <= 1 for l in raw_lines):
                    continue
                lines = [normalize(l) for l in raw_lines]
                if not lines:
                    continue
                time_label = time_labels.get(col_idx, "?")

                subject_text = normalize(lines[0], underscores=True)
                alt_text = None
                used_lines = 1
                if len(lines) > 1 and not lines[1].startswith('L') and lines[1] != 'Hybrid':
                    alt_text = normalize(subject_text + ' ' + lines[1], underscores=True)

                match = None
                for key, (code, div) in TIMETABLE_SUBJECT_MAP.items():
                    key_norm = normalize(key, underscores=True)
                    if match_subject(subject_text, key_norm):
                        match = (code, div)
                        break
                    if alt_text and (match_subject(alt_text, key_norm)):
                        match = (code, div)
                        used_lines = 2
                        break

                if not match:
                    continue

                code, div = match
                remaining = lines[used_lines:]
                room = ''
                professor = ''
                for line in remaining:
                    if line == 'Hybrid':
                        room = 'Hybrid'
                    elif re.match(r'^L\d', line):
                        if room:
                            room += ' / ' + line
                        else:
                            room = line
                    elif 'Prof' in line or 'Dr' in line:
                        if not professor:
                            professor = line
                entry = {
                    "day": day_name,
                    "time": time_label,
                    "subject": code,
                    "raw_text": '\n'.join(lines),
                    "professor": professor or "",
                    "room": room or "",
                    "div": div or "",
                }
                existing = None
                for e in timetable:
                    if (e["day"] == entry["day"] and e["time"] == entry["time"]
                            and e["subject"] == entry["subject"]
                            and e["professor"] == entry["professor"]
                            and e["room"] == entry["room"]):
                        existing = e
                        break
                if not existing:
                    timetable.append(entry)

    return timetable

def parse_food_menu(filepath):
    """Read the food menu xlsx → {week1: {day: {meal: text}}, week2: {...}}."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]

    menu = {"week1": OrderedDict(), "week2": OrderedDict()}
    current = menu["week1"]

    for row in ws.iter_rows(min_row=1, values_only=True):
        day_key = row[0]
        if not day_key:
            continue
        first = normalize(str(day_key))
        if first.lower().startswith('2nd'):
            current = menu["week2"]
            continue
        day_short = FOOD_DAY_KEYS.get(first[:3].upper())
        if day_short is None:
            continue
        entry = OrderedDict()
        for i, meal in enumerate(FOOD_MEALS):
            raw = row[1 + i] if 1 + i < len(row) else None
            text = normalize(str(raw)) if raw is not None else ""
            text = text.replace('\xa0', ' ').strip()
            entry[meal["key"]] = text
        current[day_short] = entry

    return menu

NON_STUDENT_PATTERNS = ['placements', 'faculty', 'division', 'total', 'sub total']
EXCLUDED_NAMES = ['Abhijatya Negi']
BS_DIV_OVERRIDE = {
    'Ajay Vigneshwar A': 'Div C',
    'Tina Kewlani': 'Div C',
}

def parse_students():
    # Read last year's list for roll corrections and departed-student detection
    ly_students = {}  # sap -> {name, roll, div}
    try:
        wb_ly = openpyxl.load_workbook(LAST_YEAR_LIST)
        for sheet_name in wb_ly.sheetnames:
            ws = wb_ly[sheet_name]
            div_name = sheet_name.strip()
            for row in ws.iter_rows(min_row=2, values_only=True):
                sl, name, sap, roll, gender = row
                if name and str(name).strip() and sl is not None:
                    sap = str(sap).strip() if sap else ''
                    ly_students[sap] = {
                        'name': normalize(str(name)),
                        'roll': normalize(str(roll)) if roll else '',
                        'div': f'Div {div_name[-1]}' if div_name else '',
                        'gender': normalize(str(gender)) if gender else '',
                    }
    except FileNotFoundError:
        ly_students = {}

    # Read current Trimester IV list (primary source)
    wb = openpyxl.load_workbook(STUDENT_LIST)
    ws = wb['List ']

    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    subject_cols = []
    for i, h in enumerate(headers):
        if h in SUBJECT_NAMES:
            subject_cols.append((i, h))

    students = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        if normalize(str(row[1])).lower() == 'name':
            continue

        raw_name = normalize(str(row[1]))
        roll     = normalize(str(row[5])) if row[5] else ''
        major    = normalize(str(row[7])) if row[7] else ''
        minor    = normalize(str(row[8])).replace('\xa0', '').strip() if row[8] else ''

        # Detect non-student rows (e.g. "Placements Hyd")
        if any(p in raw_name.lower() for p in NON_STUDENT_PATTERNS) and row[4]:
            name  = normalize(str(row[4]))
            email = normalize(str(row[6])) if row[6] else ''
        else:
            name  = normalize(str(row[4])) if row[4] else raw_name
            email = normalize(str(row[0])) if row[0] else ''

        # Get SAP ID — used only for roll correction from last year's list
        sap = normalize(str(row[2])) if row[2] else ''

        # Skip excluded students (e.g. left the college)
        if name in EXCLUDED_NAMES:
            continue

        # Fix roll number if last year's list has a proper HXXX format
        if sap and sap in ly_students and ly_students[sap]['roll'].startswith('H'):
            roll = ly_students[sap]['roll']

        # Get BS division from last year's list (which sheet they're in)
        bs_div = ly_students[sap]['div'] if sap and sap in ly_students else ''
        # Manual override for late-joining students not in last year's list
        if name in BS_DIV_OVERRIDE:
            bs_div = BS_DIV_OVERRIDE[name]

        # Normalize roll to Hxxx format if it starts with H
        if roll.startswith('H'):
            digits = re.sub(r'[^0-9]', '', roll)
            if digits:
                roll = 'H' + digits.zfill(3)

        gender = ly_students[sap]['gender'] if sap and sap in ly_students else ''

        subjects = []
        for col_idx, code in subject_cols:
            val = str(row[col_idx]).strip().upper() if col_idx < len(row) and row[col_idx] else 'NO'
            if val == 'YES':
                subjects.append(code)

        students.append({
            "name": name,
            "roll": roll,
            "email": email,
            "major": major,
            "minor": minor,
            "gender": gender,
            "subjects": subjects,
            "bs_div": bs_div,
        })

    # BS (Business Simulation) is a common subject — add for all
    for s in students:
        if 'BS' not in s['subjects']:
            s['subjects'].append('BS')

    return students

def parse_placements():
    """Flatten placements.json → [{name, company, date}] (in the 2-day window)."""
    placements = []
    try:
        with open(PLACEMENTS) as f:
            raw = json.load(f)
    except FileNotFoundError:
        return placements
    except Exception as e:
        print(f"⚠️ Could not parse placements: {e}")
        return placements

    today = datetime.now().date()
    window_start = today - timedelta(days=PLACEMENT_WINDOW_DAYS)

    for a in raw.get("announcements", []):
        company = str(a.get("company", "")).strip()
        date_str = str(a.get("announced_on", "")).strip()
        try:
            announced = datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            announced = None
        for name in a.get("names", []):
            roster_name = None
            if isinstance(name, dict):
                nm = str(name.get("name", "")).strip()
                roster_name = str(name.get("roster", "")).strip() or None
            else:
                nm = str(name).strip()
            if not nm:
                continue
            placements.append({
                "name": nm,
                "company": company,
                "date": date_str,
                "roster": roster_name,
            })
    return placements

def main():
    students = parse_students()
    try:
        timetable = parse_timetable(TIMETABLE)
        date_range = parse_date_range(TIMETABLE)
        week_start, week_end = get_week_iso(TIMETABLE)
    except Exception as e:
        print(f"❌ Could not parse timetable: {e}")
        timetable = []
        date_range = ""
        week_start = ""
        week_end = ""

    timetable_next = []
    date_range_next = ""
    week_start_next = ""
    week_end_next = ""
    if os.path.exists(TIMETABLE_NEXT):
        try:
            timetable_next = parse_timetable(TIMETABLE_NEXT)
            date_range_next = parse_date_range(TIMETABLE_NEXT)
            week_start_next, week_end_next = get_week_iso(TIMETABLE_NEXT)
        except Exception as e:
            print(f"⚠️ Could not parse next week's timetable: {e}")

    subjects = {}
    for code, full in SUBJECT_NAMES.items():
        students_in_subject = [s for s in students if code in s["subjects"]]
        subjects[code] = {
            "code": code,
            "name": full,
            "student_count": len(students_in_subject),
        }

    wb_now = openpyxl.load_workbook(TIMETABLE)
    ws_now = wb_now[[s for s in wb_now.sheetnames if re.match(r'\d+\.\d+\.\d{4}', s)][0]]
    time_slots = dedupe_consecutive(list(parse_time_labels_from_sheet(ws_now).values()))
    time_slots_next = []
    if os.path.exists(TIMETABLE_NEXT):
        try:
            wb_nxt = openpyxl.load_workbook(TIMETABLE_NEXT)
            ws_nxt = wb_nxt[[s for s in wb_nxt.sheetnames if re.match(r'\d+\.\d+\.\d{4}', s)][0]]
            time_slots_next = dedupe_consecutive(list(parse_time_labels_from_sheet(ws_nxt).values()))
        except Exception:
            pass

    food_menu = {}
    if os.path.exists(FOOD_MENU):
        try:
            food_menu = parse_food_menu(FOOD_MENU)
        except Exception as e:
            print(f"⚠️ Could not parse food menu: {e}")

    placements = parse_placements()

    data = {
        "students": students,
        "subjects": subjects,
        "timetable": timetable,
        "timetable_next": timetable_next,
        "time_slots": time_slots,
        "time_slots_next": time_slots_next,
        "food_menu": food_menu,
        "food_meals": FOOD_MEALS,
        "food_menu_anchor": FOOD_MENU_ANCHOR,
        "placements": placements,
        "placements_window_days": PLACEMENT_WINDOW_DAYS,
        "days": ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"],
        "date_range": date_range,
        "date_range_next": date_range_next,
        "week_start": week_start,
        "week_end": week_end,
        "week_start_next": week_start_next,
        "week_end_next": week_end_next,
    }

    # Read last_updated metadata if available (written by check_mail.py)
    last_updated_path = os.path.join(os.path.dirname(__file__), "last_updated.json")
    if os.path.exists(last_updated_path):
        try:
            with open(last_updated_path) as f:
                data["last_updated"] = json.load(f)
        except Exception:
            pass

    with open(OUTPUT, 'w') as f:
        json.dump(data, f, indent=2)
    print(f"✅ Wrote {OUTPUT}")
    print(f"   Students: {len(students)}")
    print(f"   Subjects: {len(subjects)}")
    print(f"   Timetable entries: {len(timetable)}")
    if timetable_next:
        print(f"   Next week entries: {len(timetable_next)}")
    print(f"   Food menu weeks: {len(food_menu)}")
    print(f"   Placements: {len(placements)}")

    # ─── Regenerate index.html with embedded data ───
    INDEX_HTML = "index.html"
    data_json = json.dumps(data)
    with open(INDEX_HTML) as f:
        html = f.read()

    marker = "/* DATA_INSERT_HERE */"
    if marker not in html:
        print(f"❌ Marker '{marker}' not found in {INDEX_HTML}")
        return

    before, after = html.split(marker, 1)
    # Skip everything until `var STUDENTS` — that's where real JS starts
    lines = after.split('\n')
    real_start = len(lines)
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith('var STUDENTS'):
            real_start = i
            break
    after = '\n'.join(lines[real_start:])

    init_code = '\n'
    new_html = before + marker + f"\nDATA = {data_json};" + init_code + after

    with open(INDEX_HTML, 'w') as f:
        f.write(new_html)
    print(f"✅ Updated {INDEX_HTML}")

if __name__ == '__main__':
    main()
